import string
from typing import List, Type

from transliterate import translit
from app.models import Parent, Role, Payment, Child
from sqlalchemy.orm import Session
from sqlalchemy import or_
import secrets
from flask import current_app

from flask_security import SQLAlchemyUserDatastore
from app.admin_panel.form import ChangeRoleForm

import os
from openpyxl import Workbook, load_workbook
from datetime import datetime


def make_user(db_session: Session,
              name: str, surname: str,
              patronymic: str = None,
              ) -> tuple[str, str] | None:
    if not isinstance(name, str):
        raise TypeError('Name must be strings')
    if not isinstance(surname, str):
        raise TypeError('Surname must be strings')
    if not isinstance(patronymic, str):
        raise TypeError('Patronymic must be strings')

    username = generate_username(db_session=db_session, name=name, surname=surname)
    password = generate_password()
    fs_uniquifier = secrets.token_urlsafe()

    try:
        add_user_to_db(db_session=db_session, username=username, password=password, name=name, surname=surname,
                       patronymic=patronymic, fs_uniquifier=fs_uniquifier)
        return username, password
    except Exception as e:
        db_session.rollback()
        raise Exception(e)


def generate_username(db_session: Session, name: str, surname: str) -> str:
    name_latin = translit(name, 'ru', reversed=True).lower()
    surname_latin = translit(surname, 'ru', reversed=True).lower()

    base_login = f'{name_latin[0]}.{surname_latin}'

    login = base_login
    suffix = 0

    while db_session.query(Parent).filter(Parent.username == login).scalar():
        login = f'{base_login}{suffix}'
        suffix += 1

    return login


def generate_password() -> str:
    chars = string.ascii_letters + string.digits
    return ''.join(secrets.choice(chars) for _ in range(8))


def add_user_to_db(db_session: Session,
                   username: str,
                   password: str,
                   name: str,
                   surname: str,
                   fs_uniquifier: str,
                   role: str = 'parent',
                   patronymic: str = None) -> None:
    user = Parent(username=username, name=name, surname=surname, patronymic=patronymic, fs_uniquifier=fs_uniquifier)
    user.set_password(password)

    db_session.add(user)
    db_session.flush()

    role: Role | None = db_session.query(Role).filter_by(name=role).first()
    test = db_session.query(Role).all()
    print(test, 1111111111)
    if role is None:
        raise ValueError('Role not found')

    user.roles.append(role)
    db_session.commit()


def get_user_select2(query: str, db_session: Session) -> list[Type[Parent]]:
    if not query:
        return db_session.query(Parent).order_by(Parent.surname).limit(10).all()

    return (
        db_session.query(Parent).filter(
            or_(
                Parent.name.ilike(f'%{query}%'),
                Parent.surname.ilike(f'%{query}%'),
                Parent.patronymic.ilike(f'%{query}%'),
            )
        ).order_by(Parent.surname).limit(10).all()
    )


def add_role_for_parent(db_session: Session,
                        uds: SQLAlchemyUserDatastore,
                        fs_uniquifier: str,
                        form: ChangeRoleForm) -> tuple[str, str]:
    role_name = None
    user: Parent = uds.find_user(fs_uniquifier=fs_uniquifier)

    choices = form.role.choices

    for choice in choices:
        role_id = choice[0]
        if role_id == form.role.data:
            role_name = choice[1]

    uds.add_role_to_user(user, role_name)
    db_session.commit()

    return role_name, user.name


def save_to_excel(db_session: Session):
    excel_file = 'Информация о платежах.xlsx'

    # Query to join Payment and Child tables
    data = db_session.query(
        Payment.value,
        Payment.date_pay,
        Child.name,
        Child.surname,
        Child.patronymic,
        Child.group
    ).join(
        Child, Payment.id_child == Child.id
    ).all()

    try:
        # Initialize or load the workbook
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Платежи"
            # Define headers
            ws.append(["Дата загрузки", "ФИО ребенка", "Группа", "Месяц", "Сумма"])
        else:
            wb = load_workbook(excel_file)
            ws = wb.active

        # Append each row of data
        for row in data:
            full_name = f"{row.name} {row.surname} {row.patronymic or ''}".strip()

            month_year = row.date_pay.strftime("%B %Y") if row.date_pay else ""
            ws.append([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # Current timestamp
                full_name,  # Child's full name
                row.group,  # Child's group
                month_year,  # Payment month/year
                row.value  # Payment amount
            ])

        # Save the workbook
        wb.save(excel_file)
        print(f"Данные успешно сохранены в файл: {excel_file}")

    except PermissionError:
        print(f"Ошибка: Нет доступа к файлу {excel_file}. Закройте его и попробуйте снова.")
    except Exception as e:
        print(f"Произошла ошибка при сохранении файла: {str(e)}")
